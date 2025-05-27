import openpyxl as xl
from openpyxl.chart import BarChart, Reference # for plotting bar chart


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']


    for row in range(2, sheet.max_row + 1): # ignore the first row which is the column title
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * .9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    # specify the values that we would like to plot in the bar chart
    values = Reference(sheet,
              min_row=2, max_row=sheet.max_row,
              min_col=4, max_col=4)

    # generate the bar chart
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2') # specify the location we would like to place the bar chart

    wb.save(filename)

# use the function to process our workbook
process_workbook("transactions.xlsx")
